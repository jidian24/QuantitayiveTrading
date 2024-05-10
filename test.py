import ccxt
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

def get_symbols_all_exchanges():
    # 获取所有交易所列表
    list_exchanges = ccxt.exchanges
    list_symbols = []
    df = pd.DataFrame()

    # 遍历每个交易所
    for exchange_id in list_exchanges:
        print('正在获取交易所信息---'+ exchange_id)
        try:
            # 创建交易所对象
            exchange = getattr(ccxt, exchange_id)()
            # 获取该交易所支持的交易品种
            markets = exchange.load_markets()
            symbols = list(markets.keys())
            list_symbols.extend(symbols)

            # 如果交易所正常返回数据则新增一列列名为交易所名
            df[exchange_id] = 0

            # 遍历每个交易品种
            for symbol in symbols:
                # 筛选出包含'/USDT'的交易对
                if '/USDT' in symbol:
                    # 如果交易对不在DataFrame的索引中，则添加一行并初始化为0
                    if symbol not in df.index:
                        df.loc[symbol] = 0
                    # 将支持该交易对的交易所在DataFrame中标记为1
                    df.loc[symbol, exchange_id] = 1

        except Exception as e:
            # 捕获异常并输出错误信息
            print(f"Failed to connect to or access data from exchange: {exchange_id}. Error: {e}. Skipping...")
            continue

    # 计算每一行的数值总和
    df['Sum'] = df.sum(axis=1)
    
    return df

# 调用函数获取所有交易所支持的交易品种信息
symbols_df = get_symbols_all_exchanges()

# 创建一个新的Excel文件
wb = Workbook()
ws = wb.active

# 将DataFrame写入Excel文件，并设置红色字体
for r in dataframe_to_rows(symbols_df, index=True, header=True):
    ws.append(r)

for row in ws.iter_rows(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        if cell.value == 1:
            cell.font = Font(color="FF0000")  # 设置文字颜色为红色

# 保存Excel文件
wb.save(os.path.join('..','data','all_symbols' + '.xlsx'))